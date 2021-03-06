#!/usr/bin/env python3

import re
import logging
import sqlite3
import os.path
from collections import *
import math
import json
import io

import tornado.web
import openpyxl
import tempfile
import handler
import db
import tournament
import seating
import leaderboard
import util

log = logging.getLogger('WebServer')

Player_Columns = [f.capitalize() for f in tournament.player_fields
                     if f not in ('id', 'countryid', 'flag_image')] + [
                             'Tournament']
excel_mime_type = '''
application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'''.strip()

# OpenPyXL formatting definitions
top_center_align = openpyxl.styles.Alignment(
   horizontal='center', vertical='top', wrap_text=True)
default_font = openpyxl.styles.Font()
title_font = openpyxl.styles.Font(name=default_font.name, size=14, bold=True)
column_header_font = openpyxl.styles.Font(
    name=default_font.name, size=12, bold=True)
no_border = openpyxl.styles.Border()
thin_outline = openpyxl.styles.Border(
   outline=openpyxl.styles.Side(border_style="thin")
)
paleYellowFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="FFFFCC", fill_type='solid')
paleGreenFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="DDFFDD", fill_type='solid')
paleBlueFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="DDDDFF", fill_type='solid')
blackFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="000000", fill_type='solid')

def merge_cells(sheet, row, column, height=1, width=1, 
                font=title_font, align=top_center_align, border=no_border,
                value=None, fill=None):
   if height > 1 or width > 1:
      sheet.merge_cells(
         start_row=row, start_column=column,
         end_row=row + height -1, end_column=column + width - 1)
   top_left = sheet.cell(row, column)
   top_left.alignment = align
   top_left.font = font
   top_left.border = border
   if value:
       top_left.value = value
   if fill:
       top_left.fill = fill
   return top_left

def inMergedCell(cell, sheet):
    "Test if cell is within a merged cell of the sheet"
    for mergedCellRange in sheet.merged_cells.ranges:
        if (mergedCellRange.min_col <= cell.column and
            cell.column <= mergedCellRange.max_col and
            mergedCellRange.min_row <= cell.row and
            cell.row <= mergedCellRange.max_row):
            return True
    return False

def resizeColumn(sheet, column, min_row=None, max_row=None, 
                 character_width=1.01, min_width=0, exclude_merged=True):
    base_font_size = 10
    width = min_width * base_font_size
    for tup in sheet.iter_rows(
            min_col=column, max_col=column, min_row=min_row, max_row=max_row):
        cell = tup[0]
        if cell.value and not (exclude_merged and inMergedCell(cell, sheet)):
            font_size = cell.font.sz if cell.font and cell.font.sz else 10
            width = max(width, len(str(cell.value)) * font_size)
    letter = openpyxl.utils.cell.get_column_letter(column)
    sheet.column_dimensions[letter].width = math.ceil(
        width * character_width / base_font_size)
        
def makePlayersSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet()
    sheet.title = 'Players'
    players = tournament.getPlayers(tournamentID)
    header_row = 3
    first_column = 1
    row = header_row
    columns = Player_Columns
    merge_cells(sheet, header_row - 2, first_column, 1, len(columns),
                font=title_font, border=thin_outline,
                value='{} Players'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    for i, column in enumerate(columns):
        cell = sheet.cell(row, first_column + i, value = column)
        cell.font = column_header_font
        cell.alignment = top_center_align
    for player in players:
        row += 1
        for i, f in enumerate(columns):
            cell = sheet.cell(
                row=row, column=first_column + i,
                value=tournamentName if f == 'Tournament'
                else player[f.lower()])
    for col in range(first_column, first_column + len(columns)):
        resizeColumn(sheet, col, min_row=header_row)
    return sheet

def makeSettingsSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'Settings'
    tmt_fields = [f for f in db.table_field_names('Tournaments')
                  if f not in ('Id', 'Name', 'Country', 'Owner')]
    rounds_fields = [f for f in db.table_field_names('Rounds') 
                     if f not in ('Id', 'Tournament')]
    query_fields = ['Countries.Code', 'Email'] + [
        'Tournaments.{}'.format(f) for f in tmt_fields] + [
        'Rounds.{}'.format(f) for f in rounds_fields]
    round_display_fields = [
        'Name', 'Number', 'Ordering', 'Algorithm', 'CutSize', 'Games']
    sql = """
    SELECT {} FROM Tournaments
      LEFT OUTER JOIN Countries ON Countries.Id = Tournaments.Country
      LEFT OUTER JOIN Users ON Users.Id = Tournaments.Owner
      LEFT OUTER JOIN Rounds ON Rounds.Tournament = Tournaments.Id
    WHERE Tournaments.Id = ?
    """.strip().format(','.join(query_fields))
    args = (tournamentID,)
    with db.getCur() as cur:
        cur.execute(sql, args)
        rounds = [dict(zip(map(db.fieldname, query_fields), row))
                  for row in cur.fetchall()]
    header_row = 3
    first_column = 1
    row = header_row
    merge_cells(sheet, header_row - 2, first_column, 
                1, max(2, len(round_display_fields)),
                font=title_font, border=thin_outline,
                value='{} Settings'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    top_left = first_column + (len(round_display_fields) - 2) // 2
    for field in tmt_fields + ['Code']:
        if field not in ('Name', ):
            namecell = sheet.cell(
                row, top_left, value='Country ' + field if field == 'Code' else
                'Owner ' + field if field == 'Email' else field)
            valuecell = sheet.cell(row, top_left + 1, value=rounds[0][field])
            row += 1
    row += 2
    for i, field in enumerate(round_display_fields):
        cell = sheet.cell(
            row, first_column + i,
            value = 'Seating Algorithm' if field == 'Algortihm' else
            'Cut Size' if field == 'CutSize' else
            'Round Name' if field == 'Name' else
            'Round Number' if field == 'Number' else
            field)
        cell.font = column_header_font
        cell.alignment = top_center_align
    if len(rounds) == 1 and not rounds[0]['Name']:
        row += 1
        merge_cells(sheet, row, first_column, 1,
                    len(round_display_fields), font=default_font,
                    value='No rounds defined')
    else:
        for round in rounds:
            row += 1
            for i, field in enumerate(round_display_fields):
                cell = sheet.cell(
                    row, first_column + i,
                    value = seating.ALGORITHMS[round[field] or 0].name
                    if field == 'Algorithm' else
                    seating.ORDERINGS[round[field] or 0][0]
                    if field == 'Orderings' else
                    round[field])
    for col in range(first_column, 
                     first_column + max(2, len(round_display_fields))):
        resizeColumn(sheet, col, min_row = header_row)
    return sheet

def makeScoresSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'All Scores'
    player_fields = ['Name', 'Country', 'Status']
    round_display_fields = ['Rank', 'Points', 'Penalty', 'Total']
    scoreboard, rounds = leaderboard.getTournamentScores(tournamentID)
    header_row = 3
    first_column = 1
    total_columns = len(player_fields) + len(round_display_fields) * len(rounds)
    row = header_row
    roundColorFills = [paleGreenFill, paleBlueFill]
    merge_cells(sheet, header_row - 2, first_column, 
                1, min(total_columns, 20),
                font=title_font, border=thin_outline,
                value='{} Scores'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    for i, field in enumerate(player_fields):
        cell = sheet.cell(row + 1, first_column + i, value=field)
        cell.font = column_header_font
        cell.alignment = top_center_align
    col = first_column + len(player_fields)
    color = 0
    for roundID, roundName in rounds:
        round_cell = merge_cells(
            sheet, row, col, 1, len(round_display_fields), value=roundName,
            fill=roundColorFills[color])
        for j, rfield in enumerate(round_display_fields):
            cell = sheet.cell(row + 1, col + j, value=rfield)
            cell.font = column_header_font
            cell.alignment = top_center_align
            cell.fill = roundColorFills[color]
        col += len(round_display_fields)
        color = 1 - color
    row += 1
    for player in scoreboard:
        row += 1
        for i, field in enumerate(player_fields):
            cell = sheet.cell(
                row, first_column + i, 
                value=player['type' if field == 'Status' else field.lower()])
        col = first_column + len(player_fields)
        color = 0
        for roundID, roundName in rounds:
            for j, rfield in enumerate(round_display_fields):
                cell = sheet.cell(row, col + j)
                cell.fill = roundColorFills[color]
                if roundID in player['scores']:
                    cell.value=player['scores'][roundID][
                        'score' if rfield == 'Points' else rfield.lower()]
            col += len(round_display_fields)
            color = 1 - color
        
    for col in range(first_column, first_column + total_columns):
        resizeColumn(sheet, col, min_row = header_row)
    return sheet

def makeStandingsSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'Standings'
    players, allTied = leaderboard.leaderData(tournamentID)
    fields = ([] if allTied else ['Place']) + [
        'Name', 'Country', 'Status', 'Games', 'Points', 'Penalty', 'Total']
    header_row = 3
    first_column = 1
    row = header_row
    merge_cells(sheet, header_row - 2, first_column, 1, len(fields),
                font=title_font, border=thin_outline,
                value='{} Standings'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    for i, field in enumerate(fields):
        cell = sheet.cell(
            row, first_column + i,
            value = 'Raw Points' if field == 'Points' else field)
        cell.font = column_header_font
        cell.alignment = top_center_align
    last_cut = None
    for player in players:
        row += 1
        if player['cutName'] and player['cutName'] != last_cut:
            separator = merge_cells(sheet, row, first_column, 1, len(fields),
                                    border=thin_outline, fill = blackFill)
            sheet.row_dimensions[row].height = title_font.size // 3
            row += 1
            cutname = merge_cells(sheet, row, first_column + 1,
                                  1, len(fields) - 2,
                                  border=thin_outline, font=default_font,
                                  value = 'CUT ' + player['cutName'],
                                  fill = paleYellowFill)
            for col in (first_column, first_column + len(fields) - 1):
                sheet.cell(row = row, column = col).fill = paleYellowFill
            row += 1
            last_cut = player['cutName']
        for i, field in enumerate(fields):
            cell = sheet.cell(
                row, first_column + i,
                value=player['gamesPlayed' if field == 'Games' else
                             'type' if field == 'Status' else
                             field.lower()])
    for col in range(first_column, first_column + len(fields)):
        resizeColumn(sheet, col, min_row = header_row)
    return sheet
        
def makeSeatingAndScoresSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'Seating & Scores'
    rounds = seating.getSeating(tournamentID)
    fields = (('Wind', ) if rounds and rounds[0]['winds'] else tuple()) + (
        'Name', 'Assoc.', 'Country', 'RawScore', 'Rank', 'Score', 'Penalty')
    header_row = 3
    first_column = 1
    row = header_row
    merge_cells(sheet, header_row - 2, first_column, 
                1, (len(fields) + 1) * max(1, len(rounds)),
                font=title_font, border=thin_outline,
                value='{} Seating & Scores'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    if len(rounds) == 0:
        merge_cells(sheet, row, first_column, 1, len(fields),
                    font=column_header_font,
                    value='No seating or scores found')
    roundColorFills = [paleGreenFill, paleBlueFill]
    color = 0
    rounds.sort(key=lambda r: r['round'])
    for r, round in enumerate(rounds):
        col = first_column + r * (len(fields) + 1)
        row = header_row
        merge_cells(sheet, row, col, 1, len(fields),
                    font=column_header_font,
                    value=round['name'], fill=roundColorFills[color])
        row += 1
        if len(round['tables']) == 0:
            merge_cells(sheet, row, col, 1, len(fields),
                        value='No seating or scores found',
                        fill=roundColorFills[color])
            row += 1
        last_cut = None
        for table in sorted(round['tables'], key=lambda t: t['table']):
            if table['cutName'] and table['cutName'] != last_cut:
                separator = merge_cells(sheet, row, col, 1, len(fields),
                                        border=thin_outline, fill = blackFill)
                # sheet.row_dimensions[row].height = title_font.size // 3
                row += 1
                cutname = merge_cells(sheet, row, col, 1, len(fields),
                                      border=thin_outline, font=default_font,
                                      value = 'CUT ' + table['cutName'],
                                      fill = paleYellowFill)
                row += 1
                last_cut = table['cutName']
            for c in range(col, col + len(fields)):
                sheet.cell(row, c).fill = roundColorFills[color]
            row += 1
            table_name = merge_cells(
                sheet, row, col, 1, len(fields),
                border=thin_outline, font=column_header_font,
                value = 'Table {}'.format(table['table']),
                fill=roundColorFills[color])
            row += 1
            for i, field in enumerate(fields):
                cell = sheet.cell(
                    row, col + i,
                    value='Raw Score' if field == 'RawScore' else field)
                cell.alignment = top_center_align
                cell.fill = roundColorFills[color]
            row += 1
            for wind_and_player in table['players']:
                for i, field in enumerate(fields):
                    cell = sheet.cell(
                        row, col + i,
                        value=wind_and_player['wind'] if field == 'Wind' else
                        wind_and_player['player'][
                            'association' if field == 'Assoc.' else
                            field.lower()])
                    cell.fill = roundColorFills[color]
                row += 1
            if table['unusedPoints']['rawscore']:
                extra = 1 if 'Wind' in fields else 0
                merge_cells(
                    sheet, row, col, 1, 3 + extra, font=default_font,
                    value='Unused Points', fill = roundColorFills[color])
                merge_cells(
                    sheet, row, col + 3 + extra, 1, 4, font=default_font,
                    value=table['unusedPoints']['rawscore'],
                    fill = roundColorFills[color])
                row += 1
        color = 1 - color
    # for col in range(first_column + len(fields), len(rounds) * len(fields),
    #                  len(fields) + 1):
    #     letter = openpyxl.utils.cell.get_column_letter(col)
    #     sheet.column_dimensions[letter].width = 2
    for col in range(first_column, 
                     first_column + max(1, len(rounds)) * (len(fields) + 1)):
       resizeColumn(sheet, col, min_row=header_row, min_width=2)
    return sheet
        
class DownloadTournamentSheetHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        book = openpyxl.Workbook()
        standingsSheet = makeStandingsSheet(
            book, self.tournamentid, self.tournamentname, book.active)
        seatingAndScoresSheet = makeSeatingAndScoresSheet(
            book, self.tournamentid, self.tournamentname)
        scoresSheet = makeScoresSheet(
            book, self.tournamentid, self.tournamentname)
        playersSheet = makePlayersSheet(
            book, self.tournamentid, self.tournamentname)
        settingsSheet = makeSettingsSheet(
            book, self.tournamentid, self.tournamentname)
        with tempfile.NamedTemporaryFile(
                suffix='.xlsx',
                prefix='{}_'.format(util.makeFilename(self.tournamentname))
        ) as outf:
            book.save(outf.name)
            outf.seek(0)
            self.write(outf.read())
            self.set_header('Content-type', excel_mime_type)
            self.set_header(
                'Content-Disposition',
                'attachment; filename="{}"'.format(os.path.basename(outf.name)))
            log.debug('Temporary file: {}'.format(outf.name))
        return

player_columns = [c.lower() for c in Player_Columns]
player_ID_fields = [f.capitalize() for f in db.table_field_names('Players')
                    if f not in ('Id', 'ReplacedBy')]
competition_fields = [f for f in db.table_field_names('Compete')
                      if f not in ('Id',)]

class UploadPlayersHandler(handler.BaseHandler):
    @handler.is_authenticated_ajax
    def post(self):
        tournament.initializeCountryLookup()
        encoded_players = self.get_argument('players', None)
        players = json.loads(encoded_players)
        response = {'status': 0,
                    'message': '{} candidate player records received'
                    .format(len(players))}
        rejects, nonCompete = [], []
        tournaments = defaultdict(lambda: None)
        try:
            with db.getCur() as cur:
                cur.execute('SELECT Id, Name FROM Tournaments')
                for ID, name in cur.fetchall():
                    tournaments[name] = ID
                for player in players:
                    if 'Country' in player and player['Country']:
                        player['Country'] = tournament.countryLookup[
                            player['Country']]
                    if 'Tournament' in player and player['Tournament']:
                        player['Tournament'] = tournaments[player['Tournament']]
                    if 'Type' in player:
                        player['Type'] = db.playertypecode.get(
                            player['Type'], 0)
                    ID_fields = [f for f in player 
                                 if player[f] and f in player_ID_fields]
                    if (len(ID_fields) < 2 or not player.get('Name', None) or
                        not player.get('Country', None)):
                        rejects.append(player)
                        continue
                    sql = "SELECT Id FROM Players WHERE {}".format(
                        ' AND '.join('{} = {}'.format(f, repr(player[f]))
                                     for f in ID_fields))
                    cur.execute(sql)  # Check for duplicate player
                    if len(cur.fetchall()) > 0:
                        rejects.append(player)
                        continue
                    sql = "INSERT INTO Players ({}) VALUES ({})".format(
                        ','.join(ID_fields), ','.join('?' for f in ID_fields))
                    cur.execute(sql, [player[f] for f in ID_fields])
                    player['Player'] = cur.lastrowid
                    if 'Tournament' in player and not player['Tournament']:
                        nonCompete.append(player)
                        continue
                    compete_fields = [f for f in player 
                                      if player[f] and f in competition_fields]
                    sql = "INSERT INTO Compete ({}) VALUES ({})".format(
                        ','.join(compete_fields), 
                        ','.join('?' for f in compete_fields))
                    cur.execute(sql, [player[f] for f in compete_fields])
                response['message'] = (
                    'No players uploaded.' if len(players) <= len(rejects) else
                    '{} player{} uploaded.'.format(
                        len(players) - len(rejects),
                        '' if len(players) - len(rejects) == 1 else 's'))
                if rejects:
                    response['message'] += '\n{} record{} rejected.'.format(
                        len(rejects), '' if len(rejects) == 1 else 's') + (
                            '\nReject: ' + '\nReject: '.join(
                                '{}'.format(p) for p in rejects))
                if nonCompete:
                    response['message'] += (
                        '\n{} record{} not linked to tournament.'.format(
                            len(nonCompete), 
                            '' if len(nonCompete) == 1 else 's')) + (
                            '\nNot linked: ' + '\nNot linked: '.join(
                                '{}'.format(p) for p in rejects))
        except sqlite3.DatabaseError as e:
            response['status'] = -1
            response['mesage'] = 'Error inserting player. {}'.format(e)
        return self.write(response)

class FindPlayersInSpreadsheetHandler(handler.BaseHandler):
    @handler.is_authenticated_ajax
    def post(self):
        if not ('file' in self.request.files and
                len(self.request.files['file']) > 0):
            return self.write({'status': -1,
                               'message':"Please provide a players file"})
        response = {'status': 0, 'message': 'Valid spreadsheet received'}
        content = self.request.files['file'][0]['body']
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=False, keep_vba=False,
                data_only=True, keep_links=False)
            playerLists = []
            for sheet in workbook:
                playerLists.extend(find_players_in_sheet(sheet))
            playerLists.sort(key=lambda pl: len(pl['players']),
                             reverse=True)
            response['message'] = 'Found {} player lists'.format(
                len(playerLists))
            response['playerLists'] = playerLists
        except Exception as e:
            response['status'] = -1
            response['message'] = 'Error processing spreadsheet. {}'.format(e)

        return self.write(response)

def find_players_in_sheet(sheet):
    """Return collections of players in contiguous ranges of the worksheet
    that start with a row of the expected player column headers.
    The column headers must include at least 3 of the player identification
    fields used in finding duplicate players.
    The player fields are in the rows beneath the column header row.
    The results are dictionaries of the form {'sheet': s, 'top': t,
    'players': [d]} where s is the sheet title, t is a worksheet cell like 'B7'
    for the top left corner of the contiguous range of cells, and d is
    a dictionary mapping capitalized column header names to the values for
    a particular player row.  The 'players' entry is a list of these 
    dictionaries.
    """
    result = []
    runs = {}
    last_run = None
    for rtuple in sheet.iter_rows():
        for ctuple in sheet.iter_cols(min_row=rtuple[0].row):
            cell = ctuple[0]
            head = (isinstance(cell.value, str) and
                    cell.value.lower() in player_columns)
            if head:
                extend = (last_run and last_run[1] + 1 == cell.column and
                          runs[last_run]['row'] == cell.row)
                newrun = all(cell.column < run[0] or run[1] < cell.column
                             for run in runs) 
                if extend or newrun:
                    if extend:
                        del runs[last_run]
                        last_run = (last_run[0], cell.column)
                    else:
                        last_run = (cell.column, cell.column)
                    runs[last_run] = {
                        'row': cell.row,
                        'min_column': last_run[0],
                        'max_column': last_run[1],
                        'players': {},
                        'fields': [sheet.cell(cell.row, c).value.capitalize()
                                   for c in range(last_run[0], last_run[1]+1)],
                        'max_row': cell.row
                    }
                else:
                    pass # Spurious match to column head label within players
            else:
                for run in runs if cell.value else []:
                    if (cell.row == runs[run]['max_row'] + 1 and  # contiguous
                        cell.row not in runs[run]['players'] and  # new row
                        run[0] <= cell.column and cell.column <= run[1]):
                        runs[run]['max_row'] = cell.row
                        runs[run]['players'][cell.row] = dict(
                            zip(runs[run]['fields'],
                                [sheet.cell(cell.row, c).value
                                 for c in range(runs[run]['min_column'],
                                                runs[run]['max_column'] + 1)]))
                        break
    for run in runs:
        if (len(set(runs[run]['fields']).intersection(set(player_ID_fields)))
            >= 3 and len(runs[run]['players']) > 0):
            result.append({
                'sheet': sheet.title,
                'top': sheet.cell(runs[run]['row'],
                                  runs[run]['min_column']).coordinate,
                'players': list(runs[run]['players'].values())
                })
    return result

if __name__ == '__main__':
    import os, argparse
    from pprint import *
    parser = argparse.ArgumentParser(
        description="Test player extraction from spreadsheet",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument(
        'spreadsheet', nargs='+',
        help='Spreadsheet to search for players')
    parser.add_argument(
        '-v', '--verbose', action='count', default=0,
        help='Add details on players and verbose comments')

    args = parser.parse_args()
    
    for f in args.spreadsheet:
        workbook = openpyxl.load_workbook(f)
        for sheet in workbook:
            playerLists = find_players_in_sheet(sheet)
            print(os.path.basename(f), '- worksheet', sheet.title, 'has',
                  len(playerLists), 'group{} of players'.format(
                      '' if len(playerLists) == 1 else 's'))
            for pList in playerLists:
                print('  Top cell at', pList['top'], 
                      'has {} player{} with {} attribute{}:'.format(
                          len(pList['players']), 
                          '' if len(pList['players']) == 1 else 's',
                          len(pList['players'][0]),
                          '' if len(pList['players'][0]) == 1 else 's'))
                if args.verbose > 0:
                    pprint(pList['players'])
        
